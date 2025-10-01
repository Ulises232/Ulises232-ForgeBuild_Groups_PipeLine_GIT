from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from .branch_store import Card, list_cards, upsert_card
from .catalog_queries import list_companies, list_incidence_types
from .session import current_username, get_active_user

try:  # pragma: no cover - optional dependency handled at runtime
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - fallback when openpyxl is unavailable
    Workbook = None
    load_workbook = None


__all__ = [
    "CardImportEntry",
    "CardImportSummary",
    "CardImportError",
    "load_card_entries",
    "apply_card_entries",
    "import_cards_from_file",
    "write_cards_template",
]


_SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
_SUPPORTED_CSV_EXTENSIONS = {".csv"}
_REQUIRED_KEYS = ("group", "company", "ticket", "title")
_COLUMN_ALIASES = {
    "grupo": "group",
    "group": "group",
    "empresa": "company",
    "company": "company",
    "ticket": "ticket",
    "clave": "ticket",
    "titulo": "title",
    "tituloa": "title",
    "título": "title",
    "title": "title",
    "desarrollador": "assignee",
    "developer": "assignee",
    "responsabledesarrollador": "assignee",
    "responsabledev": "assignee",
    "qa": "qa",
    "qaresponsable": "qa",
    "tipoincidencia": "incidence_type",
    "tipodeincidencia": "incidence_type",
    "incidencia": "incidence_type",
    "tipoerror": "incidence_type",
}
_TEMPLATE_HEADERS = [
    "Grupo",
    "Empresa",
    "Ticket",
    "Título",
    "Desarrollador (opcional)",
    "QA (opcional)",
    "Tipo de incidencia (opcional)",
]


class CardImportError(Exception):
    """Error genérico durante la importación de tarjetas."""


@dataclass(slots=True)
class CardImportEntry:
    """Representa una fila normalizada proveniente del archivo de importación."""

    row: int
    group_name: str
    company_name: str
    ticket_id: str
    title: str
    assignee: Optional[str] = None
    qa_assignee: Optional[str] = None
    incidence_type_name: Optional[str] = None
    incidence_type_provided: bool = False


@dataclass(slots=True)
class CardImportSummary:
    """Resultado acumulado de la importación de tarjetas."""

    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: List[Tuple[int, str]] = field(default_factory=list)

    def register_error(self, row: int, message: str) -> None:
        self.errors.append((row, message))
        self.skipped += 1


@dataclass(slots=True)
class _HeaderInfo:
    index: int
    label: str


def load_card_entries(path: Path | str) -> Tuple[List[CardImportEntry], int]:
    """Carga un archivo CSV o Excel y regresa las entradas normalizadas.

    Parameters
    ----------
    path:
        Ruta al archivo que contiene las tarjetas a importar.

    Returns
    -------
    entries, skipped_blanks:
        Una tupla con la lista de entradas válidas y el número de filas vacías
        que se omitieron.
    """

    file_path = Path(path)
    if not file_path.exists():
        raise CardImportError(f"El archivo '{file_path}' no existe.")
    suffix = file_path.suffix.lower()
    if suffix in _SUPPORTED_CSV_EXTENSIONS:
        return _load_from_csv(file_path)
    if suffix in _SUPPORTED_EXCEL_EXTENSIONS:
        return _load_from_excel(file_path)
    allowed = ", ".join(sorted(_SUPPORTED_CSV_EXTENSIONS | _SUPPORTED_EXCEL_EXTENSIONS))
    raise CardImportError(
        f"Formato no soportado para '{file_path.name}'. Usa uno de los siguientes: {allowed}."
    )


def apply_card_entries(
    entries: Sequence[CardImportEntry],
    *,
    username: Optional[str] = None,
    skipped_rows: int = 0,
    list_cards_fn: Callable[[], Iterable[Card]] | None = None,
    upsert_card_fn: Callable[[Card], Card] | None = None,
    list_companies_fn: Callable[[], Iterable] | None = None,
    list_incidence_types_fn: Callable[[], Iterable] | None = None,
) -> CardImportSummary:
    """Inserta o actualiza tarjetas a partir de entradas ya validadas."""

    summary = CardImportSummary(skipped=skipped_rows)
    if not entries:
        return summary

    username = username or _default_username()
    list_cards_fn = list_cards_fn or list_cards
    upsert_card_fn = upsert_card_fn or upsert_card
    list_companies_fn = list_companies_fn or list_companies
    list_incidence_types_fn = list_incidence_types_fn or list_incidence_types

    companies = list(list_companies_fn())
    if not companies:
        summary.register_error(0, "El catálogo de empresas está vacío. Registra empresas antes de importar.")
        return summary

    company_lookup = _build_company_lookup(companies)
    incidence_types = list(list_incidence_types_fn())
    incidence_lookup = _build_incidence_lookup(incidence_types)
    existing_cards = list(list_cards_fn())
    card_lookup = _build_card_lookup(existing_cards)

    for entry in entries:
        try:
            company = _match_company(entry, company_lookup)
        except CardImportError as exc:
            summary.register_error(entry.row, str(exc))
            continue

        ticket_key = _normalize_token(entry.ticket_id)
        group_key = _normalize_token(entry.group_name)
        card = card_lookup.get((group_key, ticket_key))

        old_key: Optional[Tuple[str, str]] = None
        if card is not None:
            old_key = (
                _normalize_token(getattr(card, "group_name", "")),
                _normalize_token(getattr(card, "ticket_id", "")),
            )

        is_new = card is None
        base_card = Card(id=None, sprint_id=None) if is_new else replace(card)
        if is_new:
            base_card.branch = ""
            base_card.status = base_card.status or "pending"
            now = int(_time_now())
            base_card.created_at = now
            base_card.updated_at = now
            base_card.created_by = username or ""
            if username:
                base_card.updated_by = username
        else:
            now = int(_time_now())
            if not base_card.status:
                base_card.status = "pending"
            base_card.updated_at = now
            if username:
                base_card.updated_by = username

        base_card.ticket_id = entry.ticket_id
        base_card.title = entry.title
        base_card.group_name = entry.group_name
        base_card.company_id = company.id
        base_card.assignee = entry.assignee or None
        base_card.qa_assignee = entry.qa_assignee or None
        if entry.incidence_type_provided:
            if entry.incidence_type_name:
                try:
                    incidence = _match_incidence_type(entry.incidence_type_name, incidence_lookup)
                except CardImportError as exc:
                    summary.register_error(entry.row, str(exc))
                    continue
                base_card.incidence_type_id = getattr(incidence, "id", None)
            else:
                base_card.incidence_type_id = None

        try:
            saved = upsert_card_fn(base_card)
        except Exception as exc:  # pragma: no cover - delegate error handling to caller/UI
            summary.register_error(entry.row, f"No se pudo guardar la tarjeta: {exc}")
            continue

        if is_new:
            summary.created += 1
        else:
            summary.updated += 1

        card_lookup[(group_key, ticket_key)] = saved
        if old_key and old_key != (group_key, ticket_key):
            card_lookup.pop(old_key, None)

    return summary


def import_cards_from_file(
    path: Path | str,
    *,
    username: Optional[str] = None,
    list_cards_fn: Callable[[], Iterable[Card]] | None = None,
    upsert_card_fn: Callable[[Card], Card] | None = None,
    list_companies_fn: Callable[[], Iterable] | None = None,
    list_incidence_types_fn: Callable[[], Iterable] | None = None,
) -> CardImportSummary:
    """Carga un archivo y aplica la importación directamente."""

    entries, skipped = load_card_entries(path)
    return apply_card_entries(
        entries,
        username=username,
        skipped_rows=skipped,
        list_cards_fn=list_cards_fn,
        upsert_card_fn=upsert_card_fn,
        list_companies_fn=list_companies_fn,
        list_incidence_types_fn=list_incidence_types_fn,
    )


def write_cards_template(path: Path | str) -> Path:
    """Genera un archivo plantilla para capturar tarjetas."""

    target = Path(path)
    if not target.suffix:
        target = target.with_suffix(".csv")
    target.parent.mkdir(parents=True, exist_ok=True)
    suffix = target.suffix.lower()

    if suffix in _SUPPORTED_CSV_EXTENSIONS:
        with target.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(_TEMPLATE_HEADERS)
    elif suffix in _SUPPORTED_EXCEL_EXTENSIONS:
        if Workbook is None:
            raise CardImportError("La librería 'openpyxl' es necesaria para generar archivos Excel.")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(_TEMPLATE_HEADERS)
        workbook.save(target)
        workbook.close()
    else:
        allowed = ", ".join(sorted(_SUPPORTED_CSV_EXTENSIONS | _SUPPORTED_EXCEL_EXTENSIONS))
        raise CardImportError(
            f"Extensión no soportada para la plantilla: usa alguno de {allowed}."
        )
    return target


def _load_from_csv(path: Path) -> Tuple[List[CardImportEntry], int]:
    rows: Optional[List[List[str]]] = None
    last_error: Optional[UnicodeDecodeError] = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            rows = _read_csv_rows(path, encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        if rows is not None:
            break

    if rows is None:
        message = (
            "No se pudo leer el archivo CSV. Guarda el archivo en UTF-8 o Latin-1 e inténtalo nuevamente."
        )
        if last_error is not None:
            raise CardImportError(message) from last_error
        raise CardImportError(message)

    if not rows:
        raise CardImportError("El archivo no contiene encabezados ni filas.")

    headers = [str(cell or "").strip() for cell in rows[0]]
    header_map = _build_header_map(headers)
    entries: List[CardImportEntry] = []
    skipped = 0

    for index, row in enumerate(rows[1:], start=2):
        values = list(row)
        entry = _entry_from_values(index, values, header_map)
        if entry is None:
            skipped += 1
            continue
        entries.append(entry)

    return entries, skipped


def _read_csv_rows(path: Path, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample or "", delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        return [list(row) for row in reader]


def _load_from_excel(path: Path) -> Tuple[List[CardImportEntry], int]:
    if load_workbook is None:
        raise CardImportError("Instala 'openpyxl' para importar archivos de Excel.")

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise CardImportError("El archivo de Excel no contiene datos.")

    headers = [str(cell or "").strip() for cell in rows[0]]
    header_map = _build_header_map(headers)
    entries: List[CardImportEntry] = []
    skipped = 0

    for index, row in enumerate(rows[1:], start=2):
        values = list(row or [])
        entry = _entry_from_values(index, values, header_map)
        if entry is None:
            skipped += 1
            continue
        entries.append(entry)

    return entries, skipped


def _build_header_map(headers: Sequence[str]) -> Dict[str, _HeaderInfo]:
    mapping: Dict[str, _HeaderInfo] = {}
    for idx, header in enumerate(headers):
        key = _map_header(header)
        if not key:
            continue
        mapping.setdefault(key, _HeaderInfo(index=idx, label=header))

    missing = [key for key in _REQUIRED_KEYS if key not in mapping]
    if missing:
        readable = ", ".join(_header_label(key) for key in missing)
        raise CardImportError(f"Faltan columnas obligatorias en el encabezado: {readable}.")
    return mapping


def _entry_from_values(
    row_index: int, values: Sequence[object], header_map: Dict[str, _HeaderInfo]
) -> Optional[CardImportEntry]:
    cleaned = [_clean_cell(values[idx]) for idx in range(len(values))]
    if all(not cell for cell in cleaned):
        return None

    def value_for(key: str) -> str:
        info = header_map.get(key)
        if not info:
            return ""
        if info.index >= len(values):
            return ""
        return _clean_cell(values[info.index])

    group = value_for("group")
    company = value_for("company")
    ticket = value_for("ticket")
    title = value_for("title")

    if not group:
        raise CardImportError(f"Fila {row_index}: el campo 'Grupo' es obligatorio.")
    if not company:
        raise CardImportError(f"Fila {row_index}: el campo 'Empresa' es obligatorio.")
    if not ticket:
        raise CardImportError(f"Fila {row_index}: el campo 'Ticket' es obligatorio.")
    if not title:
        raise CardImportError(f"Fila {row_index}: el campo 'Título' es obligatorio.")

    assignee = value_for("assignee") or None
    qa_assignee = value_for("qa") or None
    incidence_present = "incidence_type" in header_map
    incidence_value = value_for("incidence_type") if incidence_present else ""
    incidence_type = incidence_value or None

    return CardImportEntry(
        row=row_index,
        group_name=group,
        company_name=company,
        ticket_id=ticket,
        title=title,
        assignee=assignee,
        qa_assignee=qa_assignee,
        incidence_type_name=incidence_type,
        incidence_type_provided=incidence_present,
    )


def _map_header(header: str) -> Optional[str]:
    normalized = _normalize_identifier(header)
    if not normalized:
        return None
    normalized = normalized.replace("opcional", "")
    normalized = normalized.replace("responsable", "")
    normalized = normalized.replace("usuario", "")
    normalized = normalized.strip()
    return _COLUMN_ALIASES.get(normalized)


def _normalize_identifier(text: str) -> str:
    decomposed = unicodedata.normalize("NFKD", (text or "").strip().lower())
    without_marks = "".join(char for char in decomposed if not unicodedata.combining(char))
    return "".join(ch for ch in without_marks if ch.isalnum())


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _header_label(key: str) -> str:
    labels = {
        "group": "Grupo",
        "company": "Empresa",
        "ticket": "Ticket",
        "title": "Título",
        "assignee": "Desarrollador",
        "qa": "QA",
        "incidence_type": "Tipo de incidencia",
    }
    return labels.get(key, key)


def _build_company_lookup(companies: Iterable) -> Dict[Tuple[str, str], object]:
    lookup: Dict[Tuple[str, str], object] = {}
    for company in companies:
        if getattr(company, "id", None) in (None, ""):
            continue
        group_key = _normalize_token(getattr(company, "group_name", ""))
        name_key = _normalize_token(getattr(company, "name", ""))
        if not name_key:
            continue
        lookup[(group_key, name_key)] = company
    return lookup


def _build_incidence_lookup(types: Iterable) -> Dict[str, List[object]]:
    lookup: Dict[str, List[object]] = {}
    for entry in types:
        if getattr(entry, "id", None) in (None, ""):
            continue
        name_key = _normalize_token(getattr(entry, "name", ""))
        if not name_key:
            continue
        lookup.setdefault(name_key, []).append(entry)
    return lookup


def _build_card_lookup(cards: Iterable[Card]) -> Dict[Tuple[str, str], Card]:
    lookup: Dict[Tuple[str, str], Card] = {}
    for card in cards:
        ticket = _normalize_token(getattr(card, "ticket_id", ""))
        group = _normalize_token(getattr(card, "group_name", ""))
        if not ticket:
            continue
        lookup[(group, ticket)] = card
    return lookup


def _match_company(entry: CardImportEntry, lookup: Dict[Tuple[str, str], object]):
    normalized_group = _normalize_token(entry.group_name)
    normalized_name = _normalize_token(entry.company_name)
    candidate = lookup.get((normalized_group, normalized_name))
    if candidate:
        return candidate

    fallback = [value for (grp, name), value in lookup.items() if name == normalized_name]
    if not fallback:
        raise CardImportError(
            f"La empresa '{entry.company_name}' no existe en el catálogo."
        )
    if len(fallback) == 1:
        return fallback[0]
    groups = ", ".join(sorted({getattr(item, "group_name", "-") or "(sin grupo)" for item in fallback}))
    raise CardImportError(
        f"La empresa '{entry.company_name}' existe en varios grupos ({groups}); especifica el grupo correcto."
    )


def _match_incidence_type(name: str, lookup: Dict[str, List[object]]):
    normalized = _normalize_token(name)
    matches = lookup.get(normalized, [])
    if not matches:
        raise CardImportError(
            f"El tipo de incidencia '{name}' no existe en el catálogo."
        )
    if len(matches) > 1:
        raise CardImportError(
            f"El tipo de incidencia '{name}' está duplicado; verifica el catálogo."
        )
    return matches[0]


def _normalize_token(value: Optional[str]) -> str:
    if value is None:
        return ""
    decomposed = unicodedata.normalize("NFKD", str(value).strip().lower())
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def _default_username() -> str:
    username = current_username("")
    if username:
        return username
    active = get_active_user()
    if active:
        return active.username
    return ""


def _time_now() -> float:
    import time

    return time.time()

